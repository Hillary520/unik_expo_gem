# main/utils.py
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.http import HttpResponse
from datetime import datetime
from collections import defaultdict, Counter
import json
import google.generativeai as genai
import logging
import re
import base64
from io import BytesIO

from main.models import InterviewSubmission, QuestionnaireSubmission

logger = logging.getLogger(__name__)

def create_pdf_questionnaire(questionnaire):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30
    )
    story.append(Paragraph(questionnaire.name, title_style))
    story.append(Paragraph(questionnaire.description, styles['Normal']))
    story.append(Spacer(1, 20))

    # Questions
    for i, question in enumerate(questionnaire.questions_data, 1):
        # Question text
        q_text = f"{i}. {question['text']}"
        story.append(Paragraph(q_text, styles['Heading2']))
        story.append(Spacer(1, 15))

        # Answer space based on field type
        if question['fieldType'] in ['RadioGroup', 'Select']:
            # Create multiple choice format
            options = question.get('fieldOptions', [])
            for j, opt in enumerate(options):
                option_text = f"{'abcdefgh'[j]}) {opt['text']}"
                story.append(Paragraph(option_text, styles['Normal']))
            story.append(Spacer(1, 15))
            story.append(Paragraph("Answer: _______", styles['Normal']))
        
        elif question['fieldType'] == 'Switch':
            story.append(Paragraph("☐ Agree    ☐ Disagree", styles['Normal']))
        
        elif question['fieldType'] == 'Input':
            story.append(Paragraph("Answer: __________________________________________________________________", styles['Normal']))
        
        elif question['fieldType'] == 'Textarea':
            # Add multiple lines for long answers with increased space between lines
            for _ in range(5):
                story.append(Paragraph("__________________________________________________________________________________", styles['Normal']))
                story.append(Spacer(1, 25))  # Increase space between lines
        
        story.append(Spacer(2, 25))

    doc.build(story)
    buffer.seek(0)
    return buffer

def export_submissions_to_excel(item_type, item, submissions):
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Write headers
    headers = ["Respondent", "Submission Date"]
    questions = item.questions_data
    
    for question in questions:
        headers.append(question['text'])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row, submission in enumerate(submissions, 2):
        # Respondent name
        ws.cell(row=row, column=1, value=submission.respondent_name or f"Respondent {row-1}")
        
        # Submission date
        ws.cell(row=row, column=2, value=submission.submitted_at.strftime("%Y-%m-%d %H:%M"))
        
        # Answers
        for col, _ in enumerate(questions, 3):
            answer = submission.answers.get(f'q_{col-2}', '')
            ws.cell(row=row, column=col, value=answer)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={item.name}_submissions.xlsx'
    
    wb.save(response)
    return response

def generate_question_analysis(question, answers, model=None):
    try:
        generation_config = {
            "temperature": 0.7,
            "top_p": 0.95,
            "top_k": 40,
        }

        model = genai.GenerativeModel(
            model_name="gemini-1.5-flash",
            generation_config=generation_config
        )

        prompt = f"""Analyze these responses to the question: "{question['text']}"
        
        Responses:
        {json.dumps(answers, indent=2)}
        
        Provide a markdown-formatted analysis including:
        
        ## Key Findings
        - Main patterns
        - Notable trends
        
        ## Statistical Summary
        - Response distribution
        - Key metrics if applicable
        
        ## Insights
        - Important observations
        - Recommendations if applicable"""

        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        logger.error(f"Analysis generation error: {e}")
        return "Analysis generation failed"

def generate_overall_analysis(item, submissions):
    try:
        all_answers = {
            q['text']: [s.answers.get(f'q_{i+1}') for s in submissions]
            for i, q in enumerate(item.questions_data)
        }

        prompt = f"""Analyze all responses for this {item.__class__.__name__}:

        Title: {item.name}
        Description: {item.description}
        Total Responses: {len(submissions)}
        
        Questions and Responses:
        {json.dumps(all_answers, indent=2)}
        
        Provide a comprehensive markdown-formatted analysis:
        
        # Overall Analysis
        
        ## Executive Summary
        - Key findings
        - Main patterns
        
        ## Response Patterns
        - Common themes
        - Notable correlations
        
        ## Recommendations
        - Key action items
        - Areas for improvement
        
        ## Statistical Overview
        - Response rates
        - Key metrics"""

        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        logger.error(f"Overall analysis error: {e}")
        return "Overall analysis generation failed"

def parse_speed_value(speed_str):
    """Convert speed string to standardized value in Mbps"""
    try:
        speed_str = speed_str.lower().strip()
        value = float(''.join(filter(str.isdigit, speed_str)))
        
        if 'kbps' in speed_str:
            return value / 1000  # Convert to Mbps
        elif 'mbps' in speed_str:
            return value
        elif 'gbps' in speed_str:
            return value * 1000
        return value
    except:
        return 0

def prepare_chart_data(question, answers):
    if question['fieldType'] in ['RadioGroup', 'Select']:
        # Multiple choice handling remains same
        counts = defaultdict(int)
        for answer in answers:
            counts[answer] += 1
        return {
            'type': 'pie',
            'data': {
                'labels': list(counts.keys()),
                'datasets': [{
                    'data': list(counts.values()),
                    'backgroundColor': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0']
                }]
            }
        }
    elif question['fieldType'] in ['Input', 'Textarea']:
        # Check for numeric content (like speeds)
        if any(x in ' '.join(answers).lower() for x in ['mbps', 'kbps', 'gbps']):
            speeds = [parse_speed_value(ans) for ans in answers]
            return {
                'type': 'bar',
                'data': {
                    'labels': [f'Response {i+1}' for i in range(len(answers))],
                    'datasets': [{
                        'label': 'Speed (Mbps)',
                        'data': speeds,
                        'backgroundColor': '#36A2EB'
                    }]
                },
                'options': {
                    'scales': {
                        'y': {
                            'beginAtZero': True,
                            'title': {
                                'display': True,
                                'text': 'Speed (Mbps)'
                            }
                        }
                    }
                }
            }
        else:
            # Text analysis
            # Clean and tokenize text
            words = []
            for answer in answers:
                if answer:
                    words.extend(re.findall(r'\b\w+\b', answer.lower()))
            
            # Get word frequencies
            word_freq = Counter(words)
            most_common = word_freq.most_common(10)  # Top 10 words
            
            # Calculate response lengths
            lengths = [len(answer.split()) if answer else 0 for answer in answers]
            avg_length = sum(lengths) / len(lengths) if lengths else 0
            
            return {
                'type': 'text-analysis',
                'data': {
                    'wordFrequency': {
                        'type': 'bar',
                        'data': {
                            'labels': [word for word, _ in most_common],
                            'datasets': [{
                                'label': 'Word Frequency',
                                'data': [count for _, count in most_common],
                                'backgroundColor': '#36A2EB'
                            }]
                        },
                        'options': {
                            'indexAxis': 'y',
                            'scales': {
                                'x': {
                                    'title': {
                                        'display': True,
                                        'text': 'Frequency'
                                    }
                                }
                            }
                        }
                    },
                    'responseLengths': {
                        'type': 'line',
                        'data': {
                            'labels': [f'Response {i+1}' for i in range(len(answers))],
                            'datasets': [{
                                'label': 'Words per Response',
                                'data': lengths,
                                'borderColor': '#FF6384',
                                'fill': False
                            }]
                        }
                    },
                    'stats': {
                        'averageLength': avg_length,
                        'totalResponses': len(answers),
                        'uniqueWords': len(word_freq)
                    }
                }
            }
    return None

def generate_topic_analysis(topic, questionnaires, interviews):
    try:
        questionnaire_data = {}
        for q in questionnaires:
            q_submissions = QuestionnaireSubmission.objects.filter(questionnaire=q)
            questionnaire_data[q.name] = {
                question['text']: [s.answers.get(f'q_{i+1}', '') for s in q_submissions]
                for i, question in enumerate(q.questions_data)
            }

        interview_data = {}
        for i in interviews:
            i_submissions = InterviewSubmission.objects.filter(interview=i)
            interview_data[i.name] = {
                question['text']: [s.answers.get(f'q_{i+1}', '') for s in i_submissions]
                for i, question in enumerate(i.questions_data)
            }

        prompt = f"""Analyze all responses for topic "{topic.title}":

        Topic Description: {topic.description}
        
        Questionnaires: {json.dumps(questionnaire_data, indent=2)}
        
        Interviews: {json.dumps(interview_data, indent=2)}
        
        Provide a comprehensive markdown-formatted analysis:
        
        # Topic Analysis Summary
        
        ## Overview
        - Key findings across all instruments
        - Major patterns and trends
        
        ## Questionnaire Insights
        - Common themes across questionnaires
        - Notable correlations
        
        ## Interview Insights  
        - Key interview findings
        - Recurring themes
        
        ## Recommendations
        - Strategic insights
        - Areas for further investigation"""

        model = genai.GenerativeModel("gemini-1.5-flash")
        response = model.generate_content(prompt)
        return response.text
        
    except Exception as e:
        logger.error(f"Topic analysis error: {e}")
        return "Topic analysis generation failed"





